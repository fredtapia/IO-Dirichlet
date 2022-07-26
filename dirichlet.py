import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt


def extractdata(file:str, range:str):
    wb = load_workbook(file)
    dest = wb.defined_names[range].destinations
    for title, coord in dest:
        IO_table = wb[title][coord]
    return [[cell.value for cell in row] for row in IO_table]

#Mean and STDEV tables must be list of list

def getalpha(mean,stdev):
    alpha = []
    for i in range(len(mean)):  # for each row
        alpha.append([])
        for j in range(len(mean[0])):  # for each column
            alp_ij = mean[i][j] * (((mean[i][j] - (mean[i][j] ** 2.0)) / (stdev[i][j] ** 2.0))- 1.0)
            alpha[i].append(alp_ij)
    return alpha

def sumrow (data):
    return [sum([data[i][j] for j in range(len(data[i]))]) for i in range(len(data))]

def sumcolumn(data):
    return [sum([data[i][j] for i in range(len(data))]) for j in range(len(data[0]))]

def getparamkj (param:list):
    return param[-1]

def getbeta(mean, stdev):
    beta = []
    for i in range(len(mean)):  # for each row
        beta.append([])
        for j in range(len(mean[0])):  # for each column
            beta_ij = (1-mean[i][j]) *  (((mean[i][j] - (mean[i][j] ** 2.0)) / (stdev[i][j] ** 2.0))- 1.0)
            beta[i].append(beta_ij)
    return beta

def alpha_hat(mean, alpha, beta):
    alpha_kj = getparamkj(alpha)
    beta_kj = getparamkj(beta)
    alpha_hat =[]
    for i in range(len(mean)):  # for each row
        alpha_hat.append([])
        for j in range(len(mean[0])):
            alpha_hat_ij =  mean[i][j]*(alpha_kj[j]+beta_kj[j])
            alpha_hat[i].append(alpha_hat_ij)
    return alpha_hat

def getbeta_hat(alpha_hat):
    beta_hat = []
    for i in range(len(alpha_hat)):  # for each row
        beta_hat.append([])
        for j in range(len(alpha_hat[0])):
            beta_hat_ij = sumcolumn(alpha_hat)[j] - alpha_hat[i][j]
            beta_hat[i].append(beta_hat_ij)
    return beta_hat

def generateBetadist(n:int, alpha_hat, beta_hat):
    beta_dist_random = []
    beta_dist_expected = []
    for i in range(len(alpha_hat)):
        beta_dist_random.append([])
        beta_dist_expected.append([])
        for j in range(len(alpha_hat[i])):
            random_numbers = list(np.random.beta(alpha_hat[i][j],beta_hat[i][j],n))
            beta_dist_random[i].append(random_numbers)
            expected_value = np.average(random_numbers)
            beta_dist_expected[i].append(expected_value)
    return beta_dist_random

def generateGammadist(n:int, alpha):
    gamma_dist_random = []
    gamma_dist_expected = []
    for i in range(len(alpha)):
        gamma_dist_random.append([])
        gamma_dist_expected.append([])
        for j in range(len(alpha[i])):
            random_numbers = list(np.random.gamma(alpha[i][j], 1.0, n))
            gamma_dist_random[i].append(random_numbers)
            expected_value = np.average(random_numbers)
            gamma_dist_expected[i].append(expected_value)
    return gamma_dist_random

def norm_list(data):
    return [normalize(i) for i in data]
    
def generateEconLosses(A_matrix, c_vector):
    A  = np.array(A_matrix)
    I = np.identity(len(A_matrix))
    I_minus_A_inverse = np.linalg.inv(np.subtract(I,A))
    q = np.matmul(I_minus_A_inverse, c_vector)
    return list(q)

def generateAmatrices(dist_random):
    norm_dist_random = np.transpose(np.array(dist_random), (2,0,1))
    runs = []
    for i in norm_dist_random.tolist():
        norm_i = normalize(i)
        runs.append(norm_i[:-1])
    #runs = np.transpose(dist_random_array_no_VA,(2,0,1))
    return runs

def solveforq(A_matrices, c_star):
    q_list = []
    for i in A_matrices:
        A_array = np.array(i)
        I = np.identity(A_array.shape[0])
        I_minus_A_inverse = np.linalg.inv(np.subtract(I, A_array))
        q = np.matmul(I_minus_A_inverse, c_star)
        q_list.append(q.tolist())
    return q_list

def GenerateQdist(dist_random, c_star):
    runs = generateAmatrices(dist_random)
    runs_q =solveforq(runs, c_star)
    return runs_q

def normalize(data):
    normalized_list = []
    sum_column = sumcolumn(data)
    for i in range(len(data)):
        normalized_list.append([])
        for j in range(len(data[i])):
            normalized_value = data[i][j]/sum_column[j]
            normalized_list[i].append(normalized_value)
    return normalized_list

def gettype(data):
    return [type(data[i][j]) for i in range(len(data)) for j in range(len(data[0]))]

def getcstar():
    wb =load_workbook('IO_calculation.xlsx')
    dest = wb.defined_names['CSTAR'].destinations
    for title, coord in dest:
        c_star = wb[title][coord]
    return [[cell.value for cell in row] for row in c_star]

def get2018output():
    wb =load_workbook('IO_calculation.xlsx')
    dest = wb.defined_names['Output_2018'].destinations
    for title, coord in dest:
        output = wb[title][coord]
    return [[cell.value for cell in row] for row in output]

def geteconloss(q_list, output):
    econ_loss = []
    for i in range(len(q_list)):
        econ_loss_sector = []
        for j in range(len(q_list[0])):
            econ_loss_value = q_list[i][j][0] * output[j][0]
            econ_loss_sector.append(econ_loss_value)
        econ_loss.append(econ_loss_sector)
    return np.transpose(np.array(econ_loss), (1,0))

def write_to_excel(file:str, cell_range:str, data):
    wb = load_workbook(file)
    dest = wb.defined_names[cell_range].destinations
    for title, coord in dest:
        ws = wb[title][coord]
        for i in range(len(data)):
            for j in range(len(data[i])):
                ws[i][j].value = data[i][j]
    wb.save("IO table - results.xlsx")
    return "Successful!"

def getpercentile(n, data:list):
    data_array = np.array(data)
    percentile = np.percentile(data_array, n, axis=1, keepdims=True)
    return percentile

if __name__ == '__main__':
    mean = extractdata("IO table.xlsx", 'MEAN')
    std_dev = extractdata('IO table.xlsx', 'STDEV')
    sectors = extractdata('IO table.xlsx', 'SECTORS')
    alpha = getalpha(mean, std_dev)
    beta = getbeta(mean, std_dev)
    alpha_hat = alpha_hat(mean, alpha, beta)
    #print(alpha_hat)
    beta_hat = getbeta_hat(alpha_hat)
    #print(beta_hat)
    beta_distribution = generateBetadist(10000, alpha_hat, beta_hat)
    #beta_distribution2 = normalize(beta_distribution)
    #IO_means = print(beta_distribution, "\n", beta_distribution2)
    #sum = sumcolumn(beta_distribution)
    #print(sum)
    gamma_distribution = generateGammadist(1000, alpha_hat)
    #cells = write_to_excel('IO table.xlsx', 'EXPVALUE', beta_distribution2)
    #cells2 = write_to_excel('IO table - results.xlsx', 'EXPVALUE2', gamma_distribution)
    #print("Expected Value ")
    c_star = np.array(getcstar())
    output_2018 = np.array(get2018output())
    q_list_gamma = GenerateQdist(beta_distribution, c_star)
    losses = geteconloss(q_list_gamma,output_2018.tolist())
    #print(q_list_gamma)
    data_boxplot = np.transpose(np.array(q_list_gamma), (1,2,0))
    data_plot = [i[0] for i in data_boxplot.tolist()]

    loss_25th = np.multiply(getpercentile(25,data_plot),output_2018)
    loss_median = np.multiply(getpercentile(50, data_plot), output_2018)
    loss_75th = np.multiply(getpercentile(75, data_plot), output_2018)

    #print(data_boxplot)
    '''
    plt.boxplot(data_plot, showfliers=False, labels=["S%s" %(i+1) for i in range(16)])
    plt.xlabel("Economic Sectors")
    plt.ylabel("Fractional losses")
    plt.savefig("IO boxplot - 25% disruption.jpg", dpi=600)
    plt.savefig("IO boxplot - 25% disruption.pdf")
    plt.show()
    '''

    fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1, figsize=(10,15))
    #fig.set_size_inches(6.5, 8)

    bplot1 = ax1.boxplot(data_plot, showfliers=False, labels=["S%s" %(i+1) for i in range(16)])
    current_values = ax1.get_yticks()
    ax1.set_yticklabels(['{:,.0%}'.format(x) for x in current_values])
    ax1.set_title("Inoperability")
    ax1.set_ylabel("Inoperability (% loss)")
    ax1.set_xlabel("Economic Sectors")
    ax1.set_facecolor("#e8e8e8")

    bplot2 = ax2.boxplot(losses.tolist(), showfliers=False, labels=["S%s" %(i+1) for i in range(16)])
    ax2.set_title("Output Losses (million PhP)")
    ax2.set_ylabel("Losses (million Php)")
    ax2.set_xlabel("Economic Sectors")
    ax2.set_facecolor("#e8e8e8")
    current_values = ax2.get_yticks()
    ax2.set_yticklabels(['{:,.0f}'.format(x) for x in current_values])

    extent1 = ax1.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
    extent2 = ax2.get_window_extent().transformed(fig.dpi_scale_trans.inverted())
    fig.savefig('outputlosses_figure.png', bbox_inches=extent2.expanded(1.4, 1.2), dpi=600)
    fig.savefig('percentlosses_figure.png', bbox_inches=extent1.expanded(1.4, 1.2), dpi=600)

    plt.savefig("IO boxplot - 25% disruption.pdf")
    plt.show()


    # Figure

    '''    
    try_lang = [[[1, 2, 3, 4], [5, 6, 7, 8], [9, 10, 11, 12]],
                [[-1, -2, -3, -4], [-5, -6, -7, -8], [1, -10, -11, -12]],
                [[1, 2, 3, 4], [6, 7, 8, 9], [2, 3, 8, 9]],
                [[0, 0, 0, 0], [0, 0, 0, 0], [0, 0, 0, 0]]]

    Amats = generateAmatrices(try_lang)
    print(Amats)
    '''
    
    